import gspread
from google.oauth2.service_account import Credentials
import openpyxl

# ============= CONFIGURACIONES =============
SPREADSHEET_ID = "108GF6pSZ_Oc_b36cjPGm4OJf50q860Nyvk3NJBRnEGc"
SHEET_NAME = "INV ALI"

ARCHIVO_STOCK_REAL = "stock_real.xlsx"
ARCHIVO_PRINCIPAL = "stock_actual.xlsx"

COLUMNA_SKU_SHEETS = 1  # Columna A (SKU)
COLUMNA_STOCK_SHEETS = 4  # Columna D (Stock)
# ===========================================

print("Conectando a Google Sheets...")

# Autenticación
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
creds = Credentials.from_service_account_file("service_account.json", scopes=SCOPES)
client = gspread.authorize(creds)

# Abrir sheet
sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

# Leer valores
valores = sheet.get_all_values()

print("Leyendo datos desde Google Sheets...")

sku_stock = {}
for fila in valores[1:]:  # Ignorar encabezado
    if len(fila) > max(COLUMNA_SKU_SHEETS-1, COLUMNA_STOCK_SHEETS-1):
        sku = fila[COLUMNA_SKU_SHEETS-1].strip()
        stock = fila[COLUMNA_STOCK_SHEETS-1].strip()
        if sku:
            try:
                stock = int(float(stock))
            except:
                stock = 0
            sku_stock[sku] = stock

print(f"SKUs cargados desde Google Sheets: {len(sku_stock)}")

# ==================
# ACTUALIZAR stock_real.xlsx
# ==================

print("Actualizando archivo stock_real.xlsx...")

wb_real = openpyxl.load_workbook(ARCHIVO_STOCK_REAL)
ws_real = wb_real.active

for row in range(2, ws_real.max_row + 1):
    sku = str(ws_real[f"A{row}"].value).strip() if ws_real[f"A{row}"].value else None
    if sku and sku in sku_stock:
        ws_real[f"B{row}"].value = sku_stock[sku]

wb_real.save(ARCHIVO_STOCK_REAL)
print("stock_real.xlsx actualizado ✅")

# ==================
# ACTUALIZAR ARCHIVO PRINCIPAL
# ==================

print("Actualizando archivo principal...")

wb_principal = openpyxl.load_workbook(ARCHIVO_PRINCIPAL)
for nombre_hoja in wb_principal.sheetnames:
    ws = wb_principal[nombre_hoja]
    for fila in range(2, ws.max_row + 1):
        celda_sku = ws[f"G{fila}"].value
        if celda_sku:
            celda_sku = str(celda_sku).strip()
            if celda_sku in sku_stock:
                ws[f"F{fila}"].value = sku_stock[celda_sku]

wb_principal.save(ARCHIVO_PRINCIPAL.replace(".xlsx", "_actualizado.xlsx"))

print("Archivo principal actualizado ✅")
print("PROCESO COMPLETO ✅")
