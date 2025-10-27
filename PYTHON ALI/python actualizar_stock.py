import openpyxl
from openpyxl.utils import get_column_letter

# === CONFIGURACIÓN ===
archivo_principal = "stock_actual.xlsx"     # Archivo con varias hojas
archivo_referencia = "stock_real.xlsx"      # Archivo con SKUs reales
columna_sku_principal = "G"  # SKU en archivo principal
columna_stock_principal = "F"  # Stock a actualizar
columna_sku_referencia = "A"  # SKU en archivo referencia
columna_stock_referencia = "B"  # Stock real

# === CARGAR ARCHIVOS ===
print("Cargando archivos...")
wb_principal = openpyxl.load_workbook(archivo_principal)
wb_ref = openpyxl.load_workbook(archivo_referencia, data_only=True)

# Solo tomamos la primera hoja del archivo de referencia
ws_ref = wb_ref.active

# Crear diccionario de referencia SKU → stock real
print("Leyendo archivo de referencia...")
sku_stock_real = {}
for fila in ws_ref.iter_rows(min_row=2):  # Saltar encabezado
    sku = str(fila[0].value).strip() if fila[0].value is not None else None
    stock_real = fila[1].value
    if sku:
        sku_stock_real[sku] = stock_real

print(f"Referencias cargadas: {len(sku_stock_real)} SKUs")

# === ACTUALIZAR ARCHIVO PRINCIPAL ===
for nombre_hoja in wb_principal.sheetnames:
    ws = wb_principal[nombre_hoja]
    print(f"Procesando hoja: {nombre_hoja}")

    for fila in range(2, ws.max_row + 1):  # Saltar encabezado
        celda_sku = f"{columna_sku_principal}{fila}"
        celda_stock = f"{columna_stock_principal}{fila}"

        sku_val = ws[celda_sku].value
        if sku_val is None:
            continue

        sku_val = str(sku_val).strip()
        if sku_val in sku_stock_real:
            nuevo_stock = sku_stock_real[sku_val]
            # Reemplazar valor, sin alterar formato
            ws[celda_stock].value = nuevo_stock

print("Actualización completa. Guardando archivo...")

nuevo_nombre = archivo_principal.replace(".xlsx", "_actualizado.xlsx")
wb_principal.save(nuevo_nombre)

print(f"Archivo guardado como: {nuevo_nombre}")
